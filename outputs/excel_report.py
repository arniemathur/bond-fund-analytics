"""
Excel report generator for bond fund analytics.
Creates multi-sheet workbooks with metrics, charts, and formatting.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, ScatterChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule
import logging

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """
    Generates comprehensive Excel reports for bond fund analytics.
    """

    # Color palette
    COLORS = {
        'header_bg': 'FF1F4E79',  # Dark blue
        'header_font': 'FFFFFFFF',  # White
        'positive': 'FF92D050',  # Green
        'negative': 'FFFF6B6B',  # Red
        'neutral': 'FFFFF2CC',  # Light yellow
        'border': 'FFB4B4B4',  # Gray
    }

    def __init__(self):
        """Initialize the report generator."""
        self.wb = Workbook()
        self._setup_styles()

    def _setup_styles(self):
        """Set up named styles for the workbook."""
        # Header style
        self.header_style = NamedStyle(name='header_style')
        self.header_style.font = Font(bold=True, color=self.COLORS['header_font'])
        self.header_style.fill = PatternFill(
            start_color=self.COLORS['header_bg'],
            end_color=self.COLORS['header_bg'],
            fill_type='solid'
        )
        self.header_style.alignment = Alignment(horizontal='center', vertical='center')

        # Number style
        self.number_style = NamedStyle(name='number_style')
        self.number_style.number_format = '#,##0.00'
        self.number_style.alignment = Alignment(horizontal='right')

        # Percentage style
        self.pct_style = NamedStyle(name='pct_style')
        self.pct_style.number_format = '0.00%'
        self.pct_style.alignment = Alignment(horizontal='right')

        # Currency style
        self.currency_style = NamedStyle(name='currency_style')
        self.currency_style.number_format = '$#,##0.0,,"M"'
        self.currency_style.alignment = Alignment(horizontal='right')

    def _add_dataframe(
        self,
        ws,
        df: pd.DataFrame,
        start_row: int = 1,
        start_col: int = 1,
        include_index: bool = True,
        pct_columns: Optional[List[str]] = None,
        currency_columns: Optional[List[str]] = None
    ):
        """
        Add a DataFrame to a worksheet with formatting.

        Args:
            ws: Worksheet to add data to
            df: DataFrame to add
            start_row: Starting row
            start_col: Starting column
            include_index: Whether to include the index
            pct_columns: Columns to format as percentages
            currency_columns: Columns to format as currency
        """
        pct_columns = pct_columns or []
        currency_columns = currency_columns or []

        # Reset index if needed
        if include_index:
            df = df.reset_index()

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for c_idx, value in enumerate(row):
                cell = ws.cell(
                    row=start_row + r_idx,
                    column=start_col + c_idx,
                    value=value
                )

                # Apply header formatting
                if r_idx == 0:
                    cell.font = Font(bold=True, color=self.COLORS['header_font'])
                    cell.fill = PatternFill(
                        start_color=self.COLORS['header_bg'],
                        end_color=self.COLORS['header_bg'],
                        fill_type='solid'
                    )
                    cell.alignment = Alignment(horizontal='center')
                else:
                    # Apply number formatting based on column
                    col_name = df.columns[c_idx] if c_idx < len(df.columns) else ''

                    if col_name in pct_columns or 'pct' in str(col_name).lower() or 'ratio' in str(col_name).lower():
                        if isinstance(value, (int, float)) and not pd.isna(value):
                            cell.number_format = '0.00%'
                    elif col_name in currency_columns or col_name.lower() in ['aum', 'net_assets']:
                        if isinstance(value, (int, float)) and not pd.isna(value):
                            cell.number_format = '$#,##0.0,,"M"'
                    elif isinstance(value, (int, float)) and not pd.isna(value):
                        cell.number_format = '#,##0.00'

                    cell.alignment = Alignment(horizontal='right' if isinstance(value, (int, float)) else 'left')

        # Auto-adjust column widths
        for col_idx in range(len(df.columns)):
            col_letter = ws.cell(row=1, column=start_col + col_idx).column_letter
            max_length = max(
                len(str(df.columns[col_idx])),
                df.iloc[:, col_idx].astype(str).str.len().max() if len(df) > 0 else 0
            )
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

    def _add_chart(
        self,
        ws,
        chart_type: str,
        data_range: str,
        title: str,
        position: str,
        x_axis_title: str = '',
        y_axis_title: str = ''
    ):
        """Add a chart to the worksheet."""
        if chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'scatter':
            chart = ScatterChart()
        elif chart_type == 'bar':
            chart = BarChart()
        else:
            chart = LineChart()

        chart.title = title
        chart.style = 10
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = y_axis_title

        ws.add_chart(chart, position)

    def generate_summary_sheet(
        self,
        metrics_df: pd.DataFrame,
        fund_facts_df: pd.DataFrame
    ):
        """
        Generate the summary comparison sheet.

        Args:
            metrics_df: DataFrame with returns metrics
            fund_facts_df: DataFrame with fund characteristics
        """
        ws = self.wb.active
        ws.title = 'Summary'

        # Title
        ws['A1'] = 'Bond Fund Analytics Summary'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws['A2'].font = Font(italic=True, size=10)

        # Combine metrics with fund facts
        if not fund_facts_df.empty and not metrics_df.empty:
            combined = fund_facts_df.join(metrics_df, how='outer')
        elif not metrics_df.empty:
            combined = metrics_df
        else:
            combined = fund_facts_df

        if combined.empty:
            ws['A4'] = 'No data available'
            return

        # Select key columns for summary
        summary_cols = [
            'name', 'category', 'expense_ratio', 'duration', 'yield',
            'annualized_return', 'volatility', 'sharpe_ratio', 'max_drawdown',
            'alpha', 'beta', 'upside_capture', 'downside_capture'
        ]
        available_cols = [c for c in summary_cols if c in combined.columns]
        summary_df = combined[available_cols].copy()

        # Add the data
        self._add_dataframe(
            ws, summary_df, start_row=4,
            pct_columns=['expense_ratio', 'yield', 'annualized_return', 'volatility', 'max_drawdown', 'alpha']
        )

        # Add conditional formatting for Sharpe ratio
        if 'sharpe_ratio' in combined.columns:
            sharpe_col_idx = list(summary_df.columns).index('sharpe_ratio') + 2  # +2 for index and 1-based
            sharpe_range = f'{ws.cell(row=5, column=sharpe_col_idx).column_letter}5:{ws.cell(row=5+len(summary_df), column=sharpe_col_idx).column_letter}{5+len(summary_df)}'

            ws.conditional_formatting.add(
                sharpe_range,
                ColorScaleRule(
                    start_type='min', start_color='FFFF6B6B',
                    mid_type='percentile', mid_value=50, mid_color='FFFFF2CC',
                    end_type='max', end_color='FF92D050'
                )
            )

    def generate_returns_sheet(self, returns_df: pd.DataFrame):
        """
        Generate the returns time series sheet.

        Args:
            returns_df: DataFrame with monthly returns indexed by date
        """
        ws = self.wb.create_sheet('Returns')

        ws['A1'] = 'Monthly Returns Time Series'
        ws['A1'].font = Font(bold=True, size=14)

        if returns_df.empty:
            ws['A3'] = 'No returns data available'
            return

        # Format index as date strings
        returns_copy = returns_df.copy()
        returns_copy.index = returns_copy.index.strftime('%Y-%m')

        self._add_dataframe(ws, returns_copy, start_row=3)

        # Add cumulative returns section
        row_offset = len(returns_df) + 6
        ws.cell(row=row_offset, column=1, value='Cumulative Returns (Growth of $1)')
        ws.cell(row=row_offset, column=1).font = Font(bold=True, size=14)

        cumulative = (1 + returns_df).cumprod()
        cumulative.index = cumulative.index.strftime('%Y-%m')
        self._add_dataframe(ws, cumulative, start_row=row_offset + 2)

    def generate_fund_facts_sheet(self, fund_facts_df: pd.DataFrame):
        """
        Generate the fund facts detail sheet.

        Args:
            fund_facts_df: DataFrame with fund characteristics
        """
        ws = self.wb.create_sheet('Fund Facts')

        ws['A1'] = 'Fund Characteristics'
        ws['A1'].font = Font(bold=True, size=14)

        if fund_facts_df.empty:
            ws['A3'] = 'No fund facts data available'
            return

        self._add_dataframe(
            ws, fund_facts_df, start_row=3,
            pct_columns=['expense_ratio', 'yield', 'ytm', 'weighted_coupon'],
            currency_columns=['aum']
        )

    def generate_scenario_sheet(self, scenario_df: pd.DataFrame):
        """
        Generate the scenario analysis sheet.

        Args:
            scenario_df: DataFrame with scenario results
        """
        ws = self.wb.create_sheet('Scenarios')

        ws['A1'] = 'Scenario Analysis'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = 'Price impact estimates based on duration approximation'
        ws['A2'].font = Font(italic=True, size=10)

        if scenario_df.empty:
            ws['A4'] = 'No scenario data available'
            return

        self._add_dataframe(ws, scenario_df, start_row=4)

        # Add color coding for impacts
        impact_cols = [c for c in scenario_df.columns if 'rate_' in c or 'spread_' in c]
        for col in impact_cols:
            col_idx = list(scenario_df.reset_index().columns).index(col) + 1
            col_letter = ws.cell(row=1, column=col_idx).column_letter

            ws.conditional_formatting.add(
                f'{col_letter}5:{col_letter}{5+len(scenario_df)}',
                ColorScaleRule(
                    start_type='min', start_color='FFFF6B6B',  # Red for negative
                    mid_type='num', mid_value=0, mid_color='FFFFFFFF',  # White for zero
                    end_type='max', end_color='FF92D050'  # Green for positive
                )
            )

    def generate_report(
        self,
        metrics_df: pd.DataFrame,
        fund_facts_df: pd.DataFrame,
        returns_df: Optional[pd.DataFrame] = None,
        scenario_df: Optional[pd.DataFrame] = None,
        output_path: str = 'bond_fund_report.xlsx'
    ) -> str:
        """
        Generate the complete Excel report.

        Args:
            metrics_df: DataFrame with returns metrics
            fund_facts_df: DataFrame with fund characteristics
            returns_df: Optional DataFrame with returns time series
            scenario_df: Optional DataFrame with scenario results
            output_path: Path for the output file

        Returns:
            Path to the generated file
        """
        # Generate each sheet
        self.generate_summary_sheet(metrics_df, fund_facts_df)

        if returns_df is not None and not returns_df.empty:
            self.generate_returns_sheet(returns_df)

        if not fund_facts_df.empty:
            self.generate_fund_facts_sheet(fund_facts_df)

        if scenario_df is not None and not scenario_df.empty:
            self.generate_scenario_sheet(scenario_df)

        # Save workbook
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)

        logger.info(f"Report saved to {output_path}")
        return str(output_path)

    def generate_from_portfolio(self, portfolio, output_path: str = 'bond_fund_report.xlsx') -> str:
        """
        Generate report from a Portfolio object.

        Args:
            portfolio: Portfolio object with fund data
            output_path: Path for the output file

        Returns:
            Path to the generated file
        """
        return self.generate_report(
            metrics_df=portfolio.returns_metrics if portfolio.returns_metrics is not None else pd.DataFrame(),
            fund_facts_df=portfolio.get_fund_facts_df(),
            scenario_df=portfolio.scenario_results,
            output_path=output_path
        )
